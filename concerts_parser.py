"""
Полноценный парсер концертов с сайта museshow.ru/concerts/
Парсит ВСЕ концерты со страницы с автоскроллингом
С подробным логированием с ID концертов
"""

import logging
import re
from datetime import datetime
from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Настройка логирования
def setup_logging():
    """Настройка системы логирования"""
    log_format = "[%(asctime)s] %(levelname)s - %(message)s"
    
    # Создаем логгер
    logger = logging.getLogger("museshow_parser")
    logger.setLevel(logging.INFO)
    
    # Обработчик для файла
    file_handler = logging.FileHandler("parser.log", encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter(log_format))
    
    # Обработчик для консоли
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter(log_format))
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger


# Инициализация логгера
logger = setup_logging()


def get_xlsx_filename_with_timestamp():
    """Генерация имени XLSX файла с датой и временем
    
    Returns:
        str: Имя файла вида concerts_2025-10-22_00-30-15.xlsx
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"concerts_{timestamp}.xlsx"
    logger.info(f"Создан файл: {filename}")
    return filename


def init_xlsx_file(filename="concerts.xlsx"):
    """Инициализация XLSX файла с красивым форматированием"""
    logger.info(f"Инициализация XLSX файла: {filename}")
    
    # Создаем новую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Концерты"
    
    # Заголовки на русском
    headers = ["Дата концерта", "Город", "Площадка", "Программа", "Статус билетов", "Вместимость зала"]
    ws.append(headers)
    
    # Стилизация заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Устанавливаем ширину столбцов
    ws.column_dimensions['A'].width = 15  # Дата
    ws.column_dimensions['B'].width = 20  # Город
    ws.column_dimensions['C'].width = 35  # Площадка
    ws.column_dimensions['D'].width = 40  # Программа
    ws.column_dimensions['E'].width = 20  # Статус
    ws.column_dimensions['F'].width = 18  # Вместимость
    
    # Закрепляем первую строку
    ws.freeze_panes = 'A2'
    
    wb.save(filename)
    logger.info(f"XLSX файл создан с заголовками: {headers}")


def save_concert_to_xlsx(concert_data, filename="concerts.xlsx"):
    """Сохранение данных концерта в XLSX файл"""
    logger.info(f"Сохранение концерта в XLSX: {concert_data}")
    
    # Открываем существующий файл
    from openpyxl import load_workbook
    wb = load_workbook(filename)
    ws = wb.active
    
    # Добавляем строку с данными
    row_data = [
        concert_data.get("date", ""),
        concert_data.get("city", ""),
        concert_data.get("venue", ""),
        concert_data.get("program", ""),
        concert_data.get("ticket_status", ""),
        concert_data.get("available_seats", "")
    ]
    ws.append(row_data)
    
    # Применяем стили к новой строке
    row_num = ws.max_row
    border_style = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    
    for col_num in range(1, 7):
        cell = ws.cell(row=row_num, column=col_num)
        cell.border = border_style
        
        # Выравнивание
        if col_num in [1, 2, 5, 6]:  # Дата, Город, Статус, Вместимость - по центру
            cell.alignment = alignment_center
        else:  # Площадка, Программа - по левому краю
            cell.alignment = alignment_left
        
        # Цвет строки (чередование)
        if row_num % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    wb.save(filename)
    logger.info("Концерт успешно сохранен в XLSX")


def parse_available_seats(iframe, concert_idx):
    """
    Парсинг свободных мест через наведение на элементы схемы зала
    
    Args:
        iframe: Frame объект Playwright с схемой зала
        concert_idx: Номер концерта для логирования
        
    Returns:
        int: Общее количество свободных мест (сумма по всем секциям)
    """
    logger.info(f"[Концерт #{concert_idx}] Начинаем парсинг свободных мест")
    
    try:
        # Ждем загрузки схемы (уменьшено для ускорения)
        iframe.wait_for_timeout(1500)
        
        # Ищем SVG элементы (места обычно рендерятся как SVG)
        svg_elements = iframe.query_selector_all("svg circle, svg rect, svg path, svg g")
        logger.info(f"[Концерт #{concert_idx}] Найдено SVG элементов: {len(svg_elements)}")
        
        if len(svg_elements) == 0:
            logger.warning(f"[Концерт #{concert_idx}] SVG элементы не найдены")
            return 0
        
        # Словарь для хранения уникальных секций с их количеством мест
        sections_data = {}
        
        # Наводим курсор на элементы (уменьшаем до 12 для ускорения)
        for idx, element in enumerate(svg_elements[:12], 1):
            try:
                element.hover(timeout=400)
                iframe.wait_for_timeout(250)
                
                # Получаем весь текст страницы
                page_text = iframe.text_content("body")
                
                # Ищем все упоминания "Свободных мест: X"
                matches = re.findall(r'свободных мест[:\s]*(\d+)', page_text, re.IGNORECASE)
                
                if matches:
                    # Берем первое найденное значение (это текущая секция)
                    seats = int(matches[0])
                    
                    # Пытаемся найти название секции (Партер, Балкон и т.д.)
                    section_name_match = re.search(r'(Партер|Балкон|Амфитеатр|Ложа|[А-Яа-я\s]+)\s+\d+\s*[-–]\s*\d+', page_text)
                    
                    if section_name_match:
                        section_name = section_name_match.group(0)
                    else:
                        # Если название не найдено, используем индекс элемента
                        section_name = f"Секция_{idx}"
                    
                    # Добавляем секцию, если её ещё нет
                    if section_name not in sections_data:
                        sections_data[section_name] = seats
                        logger.info(f"[Концерт #{concert_idx}]   ✓ {section_name}: {seats} мест")
                
            except Exception as e:
                # Игнорируем ошибки на отдельных элементах
                pass
        
        # Суммируем все найденные места
        total_seats = sum(sections_data.values())
        
        logger.info(f"[Концерт #{concert_idx}] Найдено секций: {len(sections_data)}")
        if len(sections_data) > 0:
            logger.info(f"[Концерт #{concert_idx}] Детализация:")
            for section, seats in sections_data.items():
                logger.info(f"[Концерт #{concert_idx}]   - {section}: {seats} мест")
        
        logger.info(f"[Концерт #{concert_idx}] ✓ ИТОГО свободных мест: {total_seats}")
        return total_seats
        
    except Exception as e:
        logger.error(f"[Концерт #{concert_idx}] ✗ Ошибка при парсинге свободных мест: {e}")
        return 0


def scroll_to_load_all_concerts(page):
    """
    Скроллинг страницы для загрузки всех концертов
    
    Args:
        page: Page объект Playwright
        
    Returns:
        tuple: (количество концертов, рабочий селектор)
    """
    logger.info("Начинаем скроллинг страницы для загрузки всех концертов")
    
    # Пробуем разные селекторы
    selectors_to_try = [
        "div.elementor-loop-container > div",
        "div[data-elementor-type='loop-item']",
        "article.elementor-post",
        "div.e-loop-item",
    ]
    
    working_selector = None
    for selector in selectors_to_try:
        test_blocks = page.query_selector_all(selector)
        if len(test_blocks) > 0:
            working_selector = selector
            logger.info(f"✓ Найден рабочий селектор: '{selector}' ({len(test_blocks)} элементов)")
            break
    
    if not working_selector:
        logger.error("✗ Не найден рабочий селектор для концертов")
        return 0, None
    
    previous_count = 0
    no_change_count = 0
    scroll_iteration = 0
    
    while True:
        scroll_iteration += 1
        
        # Скроллим вниз
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(2000)  # Ждем загрузки новых элементов
        
        # Считаем количество концертов
        concert_blocks = page.query_selector_all(working_selector)
        current_count = len(concert_blocks)
        
        logger.info(f"Скроллинг #{scroll_iteration}: найдено концертов: {current_count}")
        
        # Если количество не изменилось
        if current_count == previous_count:
            no_change_count += 1
            logger.info(f"Количество концертов не изменилось ({no_change_count}/3)")
            
            # Если 3 раза подряд количество не менялось, значит все загружено
            if no_change_count >= 3:
                logger.info(f"✓ Все концерты загружены. Итого: {current_count}")
                break
        else:
            no_change_count = 0
        
        previous_count = current_count
        
        # Защита от бесконечного цикла
        if scroll_iteration > 50:
            logger.warning(f"Достигнут лимит скроллинга (50 итераций). Загружено: {current_count}")
            break
    
    return current_count, working_selector


def parse_concerts():
    """
    Парсинг ВСЕХ концертов с сайта museshow.ru с автоскроллингом
    """
    logger.info("=" * 80)
    logger.info("НАЧАЛО РАБОТЫ ПОЛНОЦЕННОГО ПАРСЕРА")
    logger.info(f"Время запуска: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 80)
    
    # Создаем XLSX файл с датой и временем
    xlsx_filename = get_xlsx_filename_with_timestamp()
    init_xlsx_file(xlsx_filename)
    
    with sync_playwright() as p:
        logger.info("Запуск Playwright")
        
        # Запуск браузера
        logger.info("Запуск браузера Chromium в headless режиме")
        browser = p.chromium.launch(headless=True)
        logger.info("Браузер успешно запущен")
        
        # Создание контекста и страницы
        logger.info("Создание контекста браузера")
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )
        logger.info("Контекст создан")
        
        logger.info("Создание новой страницы")
        page = context.new_page()
        logger.info("Страница создана")
        
        try:
            # Переход на страницу концертов
            url = "https://museshow.ru/concerts/"
            logger.info(f"Переход на URL: {url}")
            page.goto(url, wait_until="networkidle", timeout=30000)
            logger.info("Страница успешно загружена")
            
            # Ожидание загрузки контента (страница динамическая)
            logger.info("Ожидание загрузки динамического контента (5 секунд)")
            page.wait_for_timeout(5000)  # Даем время на загрузку JS
            
            # Скроллим для загрузки всех концертов
            total_concerts, working_selector = scroll_to_load_all_concerts(page)
            
            if not working_selector:
                logger.error("Не удалось найти концерты на странице")
                return
            
            # Получаем все блоки концертов используя найденный селектор
            concert_cards = page.query_selector_all(working_selector)
            logger.info(f"Начинаем парсинг {len(concert_cards)} концертов (селектор: '{working_selector}')")
            
            parsed_count = 0
            
            # Парсинг каждой карточки
            for idx, card in enumerate(concert_cards, 1):
                logger.info("=" * 60)
                logger.info(f"ОБРАБОТКА КОНЦЕРТА #{idx} (ID:{idx})")
                logger.info("=" * 60)
                
                concert_data = {
                    "date": "",
                    "city": "",
                    "venue": "",
                    "program": "",
                    "ticket_status": "",
                    "available_seats": ""
                }
                
                try:
                    # Извлечение даты
                    logger.info(f"[ID:{idx}] Извлечение даты концерта")
                    
                    # Пробуем разные селекторы для даты
                    date_selectors = [
                        "div.jet-listing-dynamic-field__content",
                        "div[class*='date']",
                        "div.elementor-widget-container",
                        "div",
                    ]
                    
                    date_text = ""
                    for date_sel in date_selectors:
                        date_elements = card.query_selector_all(date_sel)
                        if date_elements and len(date_elements) > 0:
                            date_text = date_elements[0].text_content().strip()
                            if date_text:
                                concert_data["date"] = date_text
                                logger.info(f"[ID:{idx}] ✓ Дата найдена (селектор '{date_sel}'): '{date_text}'")
                                break
                    
                    if not date_text:
                        logger.warning(f"[ID:{idx}] ✗ Дата не найдена")
                    
                    # Извлечение города и программы
                    logger.info(f"[ID:{idx}] Извлечение города и программы")
                    
                    # Пробуем найти ссылку с информацией о концерте
                    link_selectors = ["a[href*='-21-']", "a[href*='concert']", "a"]
                    link_element = None
                    for link_sel in link_selectors:
                        link_element = card.query_selector(link_sel)
                        if link_element:
                            logger.info(f"[ID:{idx}] Ссылка найдена селектором '{link_sel}'")
                            break
                    
                    if link_element:
                        full_text = link_element.text_content().strip()
                        logger.info(f"[ID:{idx}] Полный текст ссылки: '{full_text}'")
                        
                        # Извлечение города
                        city_match = re.search(r'в\s+([А-Яа-яЁё\-]+)', full_text)
                        if city_match:
                            concert_data["city"] = city_match.group(1)
                            logger.info(f"[ID:{idx}] ✓ Город найден: '{concert_data['city']}'")
                        else:
                            logger.warning(f"[ID:{idx}] ✗ Город не найден в тексте")
                        
                        # Извлечение программы
                        program_parts = full_text.split(' в ')
                        if program_parts:
                            concert_data["program"] = program_parts[0].strip()
                            logger.info(f"[ID:{idx}] ✓ Программа найдена: '{concert_data['program']}'")
                        else:
                            logger.warning(f"[ID:{idx}] ✗ Программа не найдена")
                    else:
                        logger.warning(f"[ID:{idx}] ✗ Ссылка с информацией не найдена")
                    
                    # Извлечение площадки
                    logger.info(f"[ID:{idx}] Извлечение площадки")
                    
                    # Ищем все div с классом jet-listing-dynamic-field__content
                    venue_elements = card.query_selector_all("div.jet-listing-dynamic-field__content")
                    
                    # Площадка обычно последний элемент (после даты и времени)
                    if len(venue_elements) >= 3:
                        venue_text = venue_elements[-1].text_content().strip()
                        concert_data["venue"] = venue_text
                        logger.info(f"[ID:{idx}] ✓ Площадка найдена: '{venue_text}'")
                    elif len(venue_elements) >= 2:
                        # Если элементов меньше, берем второй
                        venue_text = venue_elements[1].text_content().strip()
                        concert_data["venue"] = venue_text
                        logger.info(f"[ID:{idx}] ✓ Площадка найдена (вариант 2): '{venue_text}'")
                    else:
                        logger.warning(f"[ID:{idx}] ✗ Площадка не найдена")
                    
                    # Извлечение статуса билетов
                    logger.info(f"[ID:{idx}] Извлечение статуса билетов")
                    
                    # Пробуем разные селекторы для кнопки
                    button_selectors = [
                        "a.elementor-button span.elementor-button-text",
                        "span.elementor-button-text",
                        "a.elementor-button",
                        "a[class*='button']",
                    ]
                    
                    button = None
                    for btn_sel in button_selectors:
                        button = card.query_selector(btn_sel)
                        if button:
                            logger.info(f"[ID:{idx}] Кнопка найдена селектором '{btn_sel}'")
                            break
                    
                    if button:
                        button_text = button.text_content().strip()
                        logger.info(f"[ID:{idx}] Текст кнопки: '{button_text}'")
                        
                        if "Все билеты проданы" in button_text:
                            concert_data["ticket_status"] = "Проданы"
                            logger.info(f"[ID:{idx}] ✓ Статус: Проданы")
                        else:
                            concert_data["ticket_status"] = "Продаются"
                            logger.info(f"[ID:{idx}] ✓ Статус: Продаются")
                            
                            # Если билеты продаются, пытаемся получить вместимость
                            logger.info(f"[ID:{idx}] Попытка получить вместимость зала")
                            
                            # Пробуем найти ссылку на билеты
                            ticket_link_selectors = [
                                "a.elementor-button[href*='qtickets']",
                                "a[href*='qtickets']",
                                "a[href*='ticket']",
                                "a.elementor-button",
                            ]
                            
                            ticket_link = None
                            for tl_sel in ticket_link_selectors:
                                ticket_link = card.query_selector(tl_sel)
                                if ticket_link:
                                    href = ticket_link.get_attribute("href")
                                    if href and ("qtickets" in href or "ticket" in href):
                                        logger.info(f"[ID:{idx}] Ссылка на билеты найдена селектором '{tl_sel}'")
                                        break
                                    else:
                                        ticket_link = None
                            
                            if ticket_link:
                                ticket_url = ticket_link.get_attribute("href")
                                logger.info(f"[ID:{idx}] Ссылка на билеты: {ticket_url}")
                                
                                try:
                                    logger.info(f"[ID:{idx}] Открытие страницы билетов в новой вкладке")
                                    ticket_page = context.new_page()
                                    ticket_page.goto(ticket_url, wait_until="networkidle", timeout=15000)
                                    logger.info(f"[ID:{idx}] Страница билетов загружена")
                                    
                                    # Проверка на плашку "мероприятие прошло"
                                    logger.info(f"[ID:{idx}] Проверка на плашку 'мероприятие прошло'")
                                    event_passed = ticket_page.query_selector("div.jquery-message-container")
                                    if event_passed:
                                        message_text = event_passed.text_content().strip()
                                        logger.warning(f"[ID:{idx}] ✗ Мероприятие прошло: '{message_text}'")
                                        # Меняем статус билетов
                                        concert_data["ticket_status"] = "Мероприятие прошло"
                                        logger.info(f"[ID:{idx}] Статус изменен на 'Мероприятие прошло'")
                                        ticket_page.close()
                                        # Не прерываем, продолжаем сохранение данных
                                    else:
                                        # Получаем iframe со схемой зала
                                        logger.info(f"[ID:{idx}] Поиск iframe со схемой зала")
                                        iframe_element = ticket_page.query_selector("iframe")
                                        if iframe_element:
                                            ticket_iframe = iframe_element.content_frame()
                                            if ticket_iframe:
                                                logger.info(f"[ID:{idx}] ✓ Iframe найден")
                                                # Парсим свободные места
                                                available_seats = parse_available_seats(ticket_iframe, idx)
                                                concert_data["available_seats"] = str(available_seats) if available_seats > 0 else ""
                                            else:
                                                logger.warning(f"[ID:{idx}] ✗ Не удалось получить content_frame")
                                        else:
                                            logger.warning(f"[ID:{idx}] ✗ Iframe не найден")
                                        
                                        ticket_page.close()
                                        logger.info(f"[ID:{idx}] Страница билетов закрыта")
                                    
                                except PlaywrightTimeout:
                                    logger.error(f"[ID:{idx}] ✗ Таймаут при загрузке страницы билетов")
                                except Exception as e:
                                    logger.error(f"[ID:{idx}] ✗ Ошибка при получении вместимости: {e}")
                            else:
                                logger.warning(f"[ID:{idx}] ✗ Ссылка на билеты не найдена")
                    else:
                        logger.warning(f"[ID:{idx}] ✗ Кнопка билетов не найдена")
                    
                    # Проверка: если не удалось спарсить свободные места и статус не "Проданы"
                    if not concert_data["available_seats"] and concert_data["ticket_status"] != "Проданы":
                        logger.warning(f"[ID:{idx}] ✗ Свободные места не найдены, меняем статус")
                        concert_data["ticket_status"] = "Мероприятие прошло"
                        logger.info(f"[ID:{idx}] Статус изменен на 'Мероприятие прошло'")
                    
                    # Сохранение данных
                    logger.info(f"[ID:{idx}] Сохранение данных концерта")
                    logger.info(f"[ID:{idx}] Итоговые данные: {concert_data}")
                    save_concert_to_xlsx(concert_data, xlsx_filename)
                    
                    parsed_count += 1
                    logger.info(f"[ID:{idx}] ✓ УСПЕШНО ОБРАБОТАН")
                    logger.info(f"Прогресс: {parsed_count}/{len(concert_cards)} концертов")
                    
                except Exception as e:
                    logger.error(f"[ID:{idx}] ✗ ОШИБКА при обработке: {e}", exc_info=True)
            
            logger.info("=" * 80)
            logger.info(f"ПАРСИНГ ЗАВЕРШЕН")
            logger.info(f"Успешно обработано концертов: {parsed_count} из {len(concert_cards)}")
            logger.info(f"Данные сохранены в файл: {xlsx_filename}")
            logger.info("=" * 80)
            
        except Exception as e:
            logger.error(f"КРИТИЧЕСКАЯ ОШИБКА: {e}", exc_info=True)
        
        finally:
            logger.info("Закрытие браузера")
            browser.close()
            logger.info("Браузер закрыт")
            logger.info("Работа парсера завершена")


if __name__ == "__main__":
    logger.info("ЗАПУСК ПОЛНОЦЕННОГО ПАРСЕРА КОНЦЕРТОВ MUSESHOW.RU")
    parse_concerts()
    logger.info("Парсер завершил работу")
